import openpyxl
import os

complex = 'Реестр выполненных работ (Комплексная настройка Контур.Маркета).xlsx'
market = 'Реестр выполненных работ (Настройка Контур.Маркета).xlsx'
off_fn = 'Реестр выполненных работ (Перерегистрация ККТ без замены  ФН).xlsx'
with_fn = 'Реестр выполненных работ (Перерегистрация ККТ с заменой  ФН).xlsx'
update_kkt = 'Реестр выполненных работ (Прошивка ККТ).xlsx'
fiscal_kkt = 'Реестр выполненных работ (Регистрация и фискализация одной единицы ККТ).xlsx'
ssu_kkt = 'Реестр выполненных работ (Снятие с учета ККТ).xlsx'
blank = 'Реестр выполненных работ (Пустой бланк).xlsx'
request = 'Заявка на оказание работ по ККТ.xlsx'
request_online = 'Заявка на оказание работ по обновлению ПО ККТ.xlsx'


data_type = {
    'complex': [complex, 'complex', 'H27'],
    'market': [market, 'market', 'H25'],
    'perereg_off_fn': [off_fn, 'perereg_off_fn', 'H23'],
    'perereg_on_fn': [with_fn, 'perereg_on_fn', 'H23'],
    'ssu': [ssu_kkt, 'ssu', 'H21'],
    'update': [update_kkt, 'update', 'H24'],
    'blank': [blank, 'blank', 'H23']
}

data_path = f'{os.path.abspath("")}' + '\\data\\'


def fiscal_and_reg(spec_name, **kwargs):
    open_file = openpyxl.load_workbook(f'{data_path}{fiscal_kkt}')
    data = open_file['fiscal']
    data['H24'] = spec_name
    for new in kwargs.items():
        data[new[0]] = new[1]
    open_file.save(f'{data_path}{fiscal_kkt}')


def complex_settings(spec_name, filename, *args, **kwargs):
    open_file = openpyxl.load_workbook(f'{data_path}' + f'{filename}')
    data = open_file[args[1]]
    data[args[2]] = spec_name
    for new in kwargs.items():
        if new[0] == 'H5':
            continue
        data[new[0]] = new[1]
    open_file.save(f'{data_path}' + f'{filename}')


def request_for_work(filename, **kwargs):
    request_file = openpyxl.load_workbook(f'{data_path}' + f'{filename}')
    data = request_file['form']
    for new in kwargs.items():
        data[new[0]] = new[1]
    request_file.save(f'{data_path}' + f'{filename}')


def make_mail(filename):
    file_path = f'{os.path.abspath("")}' + f'\\data\\{filename}'
    text_file = open(f'{data_path}text_mail.txt', encoding='utf_8_sig')
    body = []
    for line in text_file:
        body.append(line)
    os.system(f'start outlook.exe /a "{file_path}" /c ipm.note /m " ?{body[0][:-1]}&{body[1]}"')
